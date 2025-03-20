import openpyxl
import os
import re

from openpyxl.styles.builtins import output

write_2_cs_root = "../Assets/Scripts/Tables/"
def excel_name2cs_class_name(input_str):
    return re.sub("(.xlsx)","Table", input_str)

def excel_name2cs_path(excel_name):
    class_name = excel_name[0].upper() + excel_name[1:] + ".cs"
    class_name = excel_name2cs_class_name(class_name)
    return os.path.join(write_2_cs_root, class_name)

directory = "./"
for root, dirs, files in os.walk(directory):
    for file in files:
        if ".xlsx" in file or ".xls" in file and not "~" in file:
            file_path = os.path.join(root, file)
            print(file_path)
            output_path = excel_name2cs_path(file_path[2:])
            print(output_path)
            print("Processing... file_path = " + file_path)  # 打印完整路径
            # 打开Excel文件
            workbook = openpyxl.load_workbook(file_path)
            # 获取Sheet1工作表 所有单元格装入内存
            sheet_name = "Sheet1"
            tab_name_list = []
            tab_type_list = []
            data = {}
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                row_counter = 1
                for row in sheet.iter_rows():
                    if row_counter == 1:
                        for cell in row:
                            tab_name_list.append(cell.value)
                        for tab_name in tab_name_list:
                            data[tab_name] = []
                    elif row_counter == 2:
                        print("skip")
                    elif row_counter == 3:
                        for cell in row:
                            tab_type_list.append(cell.value)
                    else:
                        for cell in row:
                            tab_name = tab_name_list[cell.col_idx - 1]
                            data[tab_name].append(cell.value)
                    row_counter = row_counter + 1
            else:
                print(f"工作表 {sheet_name} 不存在！")
                workbook.close()
                exit()
            # 关闭工作簿
            workbook.close()

            replaced_article = []
            try:
                # 打开文件
                with open(output_path, 'r', encoding='utf-8') as out:
                    # 逐行读取并打印文件内容
                    article = []
                    generated_content = []
                    # List<int> numbers = new List<int>(new int[] { 1, 2, 3, 4, 5 });
                    generated_line = ""
                    for i in range(len(tab_name_list)):
                        tab_name = tab_name_list[i]
                        tab_type = tab_type_list[i]
                        generated_line = " List<" + tab_type + "> " + tab_name + "s = new List<" + tab_type +">(new " + tab_type + "[] {"
                        line_p2 = ""
                        for i in range(len(data[tab_name])):
                            dataValue = data[tab_name][i]
                            if tab_type == "string":
                                dataValue = "\"" + dataValue + "\""
                            if i != 0:
                                line_p2 = line_p2 + ", "
                            line_p2 = line_p2 + str(dataValue)

                        generated_line = generated_line + line_p2 + " });\n"
                        generated_content.append(generated_line)

                    replace_start_idx = 1
                    replace_end_idx = 1
                    for line_number, line in enumerate(out, start=1):
                        article.append(line.strip())
                        if "#region auto_generate_start" in line.strip():
                            replace_start_idx = line_number - 1
                        elif "#endregion auto_generate_end" in line.strip():
                            replace_end_idx = line_number - 1

                    for i in range(len(article)):
                        l = article[i]
                        if i <= replace_start_idx:
                            replaced_article.append(l)

                    for i in range(len(generated_content)):
                        l = generated_content[i]
                        replaced_article.append(l)
                    replaced_article.append("\n")
                    for i in range(len(article)):
                        l = article[i]
                        print(l)
                        if i >= replace_end_idx:
                            replaced_article.append(l)

            except FileNotFoundError:
                print(f"文件 {output_path} 未找到。")
            except Exception as e:
                print(f"读取文件时发生错误：{e}")

            try:
                with open(output_path, 'w', encoding='utf-8') as out:
                    for l in replaced_article:
                        out.write(l + "\n")
            except FileNotFoundError:
                print(f"文件 {output_path} 写入阶段找不到。")
            except Exception as err:
                print(f"写入目标文件时发生错误：{err}")
